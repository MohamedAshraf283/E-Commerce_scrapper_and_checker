from __future__ import annotations

import os
import shutil
import tempfile
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .models import InputRow, ProductRow, StoreResult
from .utils import ensure_parent_dir, safe_str


LINK_COLUMN_CANDIDATES = [
    "link", "url", "store_url", "store link", "رابط", "رابط المتجر", "المتجر", "store",
]

STORE_COLUMNS = [
    "#", "Excel Row", "Input URL", "Final URL", "Platform", "Store Name",
    "Status", "Status Reason", "HTTP/Error", "Emails", "Phones", "WhatsApp",
    "Socials", "Address", "Categories Found", "Products Found",
    "Started At", "Finished At",
]

PRODUCT_COLUMNS = [
    "Store #", "Store URL", "Category URL", "Product URL", "Product Name",
    "Price", "Currency", "Image", "Availability", "Raw Text",
]

SUMMARY_COLUMNS = ["Metric", "Value"]


def normalize_header(value: Any) -> str:
    return safe_str(value)


def headers_from_sheet(ws) -> list[str]:
    headers = []
    for cell in ws[1]:
        value = normalize_header(cell.value)
        headers.append(value if value else f"Column {cell.column}")
    return headers


def detect_link_column(headers: list[str]) -> str:
    lowered = {h.strip().lower(): h for h in headers}
    for candidate in LINK_COLUMN_CANDIDATES:
        if candidate.lower() in lowered:
            return lowered[candidate.lower()]

    for h in headers:
        low = h.lower()
        if "url" in low or "link" in low or "رابط" in low:
            return h
    return headers[0] if headers else "link"


def read_input_rows(path: str, link_column: str = "", sheet_name: str | None = None) -> tuple[list[str], list[InputRow], str]:
    wb = load_workbook(path, data_only=False)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    headers = headers_from_sheet(ws)
    chosen_link_col = link_column if link_column in headers else detect_link_column(headers)

    if chosen_link_col not in headers:
        raise ValueError(f"عمود الروابط غير موجود: {chosen_link_col}")

    link_index = headers.index(chosen_link_col)

    rows: list[InputRow] = []

    for idx, row in enumerate(ws.iter_rows(min_row=2), start=0):
        values = []

        for i in range(len(headers)):
            cell = row[i] if i < len(row) else None
            value = cell.value if cell else ""

            # لو الخلية Hyperlink، نستخدم الرابط الحقيقي بدل النص الظاهر
            if i == link_index and cell is not None and cell.hyperlink and cell.hyperlink.target:
                target = safe_str(cell.hyperlink.target)
                shown = safe_str(value)
                value = target if target else shown

            values.append(value)

        data = {headers[i]: values[i] if i < len(values) else "" for i in range(len(headers))}
        link_value = safe_str(data.get(chosen_link_col, "")).strip()

        # لا تضف الصف أصلاً لو عمود link فاضي
        if not link_value:
            continue

        rows.append(InputRow(index=len(rows), excel_row=idx + 2, data=data))

    return headers, rows, chosen_link_col


def preview_rows(path: str, limit: int = 100) -> tuple[list[str], list[list[Any]], str]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = headers_from_sheet(ws)
    rows = []
    for row in ws.iter_rows(min_row=2, max_row=limit + 1, values_only=True):
        if any(safe_str(v) for v in row):
            rows.append([safe_str(v) for v in row[: len(headers)]])
    return headers, rows, detect_link_column(headers)


def _style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for column_cells in ws.columns:
        max_len = 10
        col_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells[:200]:
            value = safe_str(cell.value)
            if value:
                max_len = max(max_len, min(len(value), 70))
        ws.column_dimensions[col_letter].width = max_len + 2


def _atomic_save(wb: Workbook, path: str) -> None:
    import time
    import datetime

    ensure_parent_dir(path)

    directory = os.path.dirname(os.path.abspath(path)) or "."
    filename = os.path.basename(path)
    base, ext = os.path.splitext(filename)
    ext = ext or ".xlsx"

    fd, tmp_path = tempfile.mkstemp(
        prefix=f".{base}.",
        suffix=ext,
        dir=directory
    )
    os.close(fd)

    try:
        wb.save(tmp_path)

        last_error = None

        for _ in range(8):
            try:
                os.replace(tmp_path, path)
                return
            except PermissionError as exc:
                last_error = exc
                time.sleep(0.75)
            except FileExistsError as exc:
                last_error = exc
                try:
                    os.remove(path)
                except Exception:
                    time.sleep(0.75)

        stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        fallback_path = os.path.join(directory, f"{base}_{stamp}{ext}")

        os.replace(tmp_path, fallback_path)

        print(
            f"WARNING: ملف الإخراج مقفول ولا يمكن الكتابة عليه:\n"
            f"{path}\n"
            f"تم حفظ نسخة بديلة هنا:\n"
            f"{fallback_path}\n"
            f"Original error: {last_error}"
        )

    finally:
        if os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except Exception:
                pass

def make_summary(results: list[StoreResult]) -> list[tuple[str, Any]]:
    total = len(results)
    by_status: dict[str, int] = {}
    by_platform: dict[str, int] = {}
    errors = 0
    success_products = 0

    for result in results:
        status = result.status or "غير مكتمل"
        by_status[status] = by_status.get(status, 0) + 1
        platform = result.platform or "غير معروف"
        by_platform[platform] = by_platform.get(platform, 0) + 1
        if status not in ["يعمل", "تم السحب"]:
            errors += 1
        if result.products_found > 0:
            success_products += 1

    rows: list[tuple[str, Any]] = [
        ("Total stores", total),
        ("Stores with products", success_products),
        ("Stores with errors / not working", errors),
    ]
    for status, count in sorted(by_status.items()):
        rows.append((f"Status: {status}", count))
    for platform, count in sorted(by_platform.items()):
        rows.append((f"Platform: {platform}", count))
    return rows


def write_output(path: str, results: list[StoreResult]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Stores"
    ws.append(STORE_COLUMNS)

    products: list[ProductRow] = []

    for result in results:
        ws.append([
            result.index + 1,
            result.excel_row,
            result.input_url,
            result.final_url,
            result.platform,
            result.store_name,
            result.status,
            result.status_reason,
            result.http_or_error,
            result.emails,
            result.phones,
            result.whatsapp,
            result.socials,
            result.address,
            result.categories_found,
            result.products_found,
            result.started_at,
            result.finished_at,
        ])
        products.extend(result.products or [])

    _style_sheet(ws)

    product_ws = wb.create_sheet("Products")
    product_ws.append(PRODUCT_COLUMNS)
    for product in products:
        product_ws.append([
            product.store_index + 1,
            product.store_url,
            product.category_url,
            product.product_url,
            product.name,
            product.price,
            product.currency,
            product.image,
            product.availability,
            product.raw_text,
        ])
    _style_sheet(product_ws)

    summary_ws = wb.create_sheet("Summary")
    summary_ws.append(SUMMARY_COLUMNS)
    for row in make_summary(results):
        summary_ws.append(list(row))
    _style_sheet(summary_ws)

    _atomic_save(wb, path)


def get_sheet_names(path: str) -> list[str]:
    wb = load_workbook(path, read_only=True)
    return list(wb.sheetnames)


def sort_excel(input_path: str, output_path: str, sheet_name: str, column_name: str, numeric: bool = False, descending: bool = False) -> None:
    wb = load_workbook(input_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"الشيت غير موجود: {sheet_name}")
    ws = wb[sheet_name]
    headers = headers_from_sheet(ws)
    if column_name not in headers:
        raise ValueError(f"العمود غير موجود: {column_name}")

    col_index = headers.index(column_name)
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    def key_func(row):
        value = row[col_index] if col_index < len(row) else None
        if value is None:
            return (1, "")
        if numeric:
            try:
                return (0, float(str(value).replace(",", "").strip()))
            except Exception:
                return (1, 0)
        return (0, str(value).strip().lower())

    rows.sort(key=key_func, reverse=descending)

    for row_number in range(2, ws.max_row + 1):
        ws.delete_rows(2)

    for row in rows:
        ws.append(list(row))

    _style_sheet(ws)
    _atomic_save(wb, output_path)
