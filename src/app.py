from __future__ import annotations

import asyncio
import json
import os
import queue
import threading
import traceback
from dataclasses import asdict
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
from openpyxl import load_workbook
from .excel_io import detect_link_column, get_sheet_names, preview_rows, sort_excel
from .models import ScrapeSettings, StoreResult
from .scraper import parse_selectors_json, run_scraper
from .utils import clamp_float, clamp_int, safe_str

APP_TITLE = "Salla / Zid Desktop Scraper"
DEFAULT_SELECTORS_JSON = """{
  "category_links": "",
  "product_card": "",
  "product_name": "",
  "product_price": "",
  "product_url": "",
  "product_image": ""
}"""

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("1240x820")
        self.minsize(1050, 720)

        self.ui_queue: queue.Queue = queue.Queue()
        self.worker_thread: threading.Thread | None = None
        self.stop_event = threading.Event()
        self.headers: list[str] = []
        self.results: list[StoreResult] = []

        self._init_vars()
        self._build_ui()
        self._poll_queue()

    def _init_vars(self):
        self.input_excel_var = tk.StringVar()
        self.output_excel_var = tk.StringVar(value=str(Path.cwd() / "output" / "salla_zid_results.xlsx"))
        self.browser_exe_var = tk.StringVar(
            value=r"C:\Program Files\BraveSoftware\Brave-Browser\Application\brave.exe"
        )

        self.user_data_dir_var = tk.StringVar(
            value=r"C:\Users\Mohamed Ashraf\AppData\Local\BraveSoftware\Brave-Browser\User Data\Default"
        )
        self.profile_dir_var = tk.StringVar(value="Default")
        self.link_column_var = tk.StringVar(value="link")
        self.category_column_var = tk.StringVar(value="")

        self.mode_status_var = tk.BooleanVar(value=True)
        self.mode_contacts_var = tk.BooleanVar(value=True)
        self.mode_products_var = tk.BooleanVar(value=True)
        self.headless_var = tk.BooleanVar(value=False)

        self.concurrency_var = tk.StringVar(value="2")
        self.timeout_var = tk.StringVar(value="30000")
        self.save_every_var = tk.StringVar(value="5")
        self.restart_every_var = tk.StringVar(value="100")
        self.min_delay_var = tk.StringVar(value="1.0")
        self.max_delay_var = tk.StringVar(value="3.0")
        self.human_wait_var = tk.StringVar(value="60")
        self.human_retries_var = tk.StringVar(value="1")
        self.max_categories_var = tk.StringVar(value="20")
        self.max_category_pages_var = tk.StringVar(value="3")
        self.max_products_var = tk.StringVar(value="200")
        self.selectors_text: tk.Text | None = None
        self.guide_text: tk.Text | None = None

        self.done_var = tk.StringVar(value="0")
        self.remaining_var = tk.StringVar(value="0")
        self.total_var = tk.StringVar(value="0")
        self.status_var = tk.StringVar(value="جاهز")

        self.sort_input_var = tk.StringVar()
        self.sort_output_var = tk.StringVar(value=str(Path.cwd() / "output" / "sorted.xlsx"))
        self.sort_sheet_var = tk.StringVar()
        self.sort_column_var = tk.StringVar()
        self.sort_numeric_var = tk.BooleanVar(value=False)
        self.sort_desc_var = tk.BooleanVar(value=False)

    def _build_ui(self):
        self.columnconfigure(0, weight=1)
        self.rowconfigure(0, weight=1)

        self.notebook = ttk.Notebook(self)
        self.notebook.grid(row=0, column=0, sticky="nsew", padx=8, pady=8)

        self.settings_tab = ttk.Frame(self.notebook)
        self.advanced_tab = ttk.Frame(self.notebook)
        self.sheet_tab = ttk.Frame(self.notebook)
        self.run_tab = ttk.Frame(self.notebook)
        self.sort_tab = ttk.Frame(self.notebook)

        self.notebook.add(self.settings_tab, text="الإعدادات")
        self.notebook.add(self.advanced_tab, text="Selectors والتصنيفات")
        self.notebook.add(self.sheet_tab, text="عرض الشيت")
        self.notebook.add(self.run_tab, text="التشغيل والتقرير")
        self.notebook.add(self.sort_tab, text="ترتيب Excel")

        self._build_settings_tab()
        self._build_advanced_tab()
        self._build_sheet_tab()
        self._build_run_tab()
        self._build_sort_tab()
        self._build_footer()

    def _help(self, title: str, text: str):
        messagebox.showinfo(title, text)

    def _field(self, parent, row: int, label: str, variable: tk.StringVar, help_text: str, browse: str | None = None, width: int = 55):
        ttk.Label(parent, text=label).grid(row=row, column=0, sticky="w", padx=6, pady=5)
        entry = ttk.Entry(parent, textvariable=variable, width=width)
        entry.grid(row=row, column=1, sticky="ew", padx=6, pady=5)
        if browse == "open_file":
            ttk.Button(parent, text="اختيار", command=lambda: self._browse_open_file(variable)).grid(row=row, column=2, padx=3, pady=5)
        elif browse == "save_file":
            ttk.Button(parent, text="حفظ باسم", command=lambda: self._browse_save_file(variable)).grid(row=row, column=2, padx=3, pady=5)
        elif browse == "dir":
            ttk.Button(parent, text="مجلد", command=lambda: self._browse_dir(variable)).grid(row=row, column=2, padx=3, pady=5)
        else:
            ttk.Label(parent, text="").grid(row=row, column=2, padx=3, pady=5)
        ttk.Button(parent, text="؟", width=3, command=lambda: self._help(label, help_text)).grid(row=row, column=3, padx=3, pady=5)
        parent.columnconfigure(1, weight=1)
        return entry

    def _build_settings_tab(self):
        self.settings_tab.columnconfigure(0, weight=1)
        self.settings_tab.rowconfigure(3, weight=1)
        frame = ttk.LabelFrame(self.settings_tab, text="ملفات التشغيل وجلسة المتصفح")
        frame.grid(row=0, column=0, sticky="ew", padx=10, pady=10)

        self._field(frame, 0, "ملف Excel المصدر", self.input_excel_var,
                    "ملف xlsx الذي يحتوي على روابط المتاجر. بعد اختياره اضغط زر قراءة الشيت في تبويب عرض الشيت.",
                    "open_file")
        self._field(frame, 1, "ملف Excel الناتج", self.output_excel_var,
                    "الملف الذي سيتم إنشاء النتائج فيه. سيتم إنشاء Sheets: Stores و Products و Summary.",
                    "save_file")
        self._field(frame, 2, "مسار المتصفح exe", self.browser_exe_var,
                    "اختياري. مثال Brave: C:\\Program Files\\BraveSoftware\\Brave-Browser\\Application\\brave.exe. اتركه فارغًا لاستخدام Chromium الخاص بـ Playwright.",
                    "open_file")
        self._field(frame, 3, "مجلد User Data", self.user_data_dir_var,
                    "المجلد الجذر للبروفايلات، مثل: C:\\Users\\User\\AppData\\Local\\BraveSoftware\\Brave-Browser\\User Data. لا تضع \\Default هنا.",
                    "dir")
        self._field(frame, 4, "اسم Profile", self.profile_dir_var,
                    "اسم البروفايل داخل User Data، غالبًا Default أو Profile 1.",
                    None, width=25)

        sheet_frame = ttk.LabelFrame(self.settings_tab, text="الأعمدة والمودات")
        sheet_frame.grid(row=1, column=0, sticky="ew", padx=10, pady=10)
        sheet_frame.columnconfigure(1, weight=1)

        ttk.Label(sheet_frame, text="عمود الروابط").grid(row=0, column=0, sticky="w", padx=6, pady=5)
        self.link_column_combo = ttk.Combobox(sheet_frame, textvariable=self.link_column_var, values=[], state="normal")
        self.link_column_combo.grid(row=0, column=1, sticky="ew", padx=6, pady=5)
        ttk.Button(sheet_frame, text="؟", width=3, command=lambda: self._help(
            "عمود الروابط", "اكتب أو اختر اسم العمود الذي يحتوي على رابط المتجر. البرنامج يحاول اكتشافه تلقائيًا من أسماء مثل link/url/رابط."
        )).grid(row=0, column=2, padx=3, pady=5)

        modes = ttk.Frame(sheet_frame)
        modes.grid(row=1, column=1, sticky="w", padx=6, pady=5)
        ttk.Checkbutton(modes, text="فحص الحالة", variable=self.mode_status_var).pack(side="left", padx=8)
        ttk.Checkbutton(modes, text="بيانات التواصل", variable=self.mode_contacts_var).pack(side="left", padx=8)
        ttk.Checkbutton(modes, text="المنتجات", variable=self.mode_products_var).pack(side="left", padx=8)
        ttk.Button(sheet_frame, text="شرح المودات", command=lambda: self._help(
            "المودات",
            "فحص الحالة: يتأكد من أن المتجر يعمل أو لا.\n"
            "بيانات التواصل: يحلل الهيدر والفوتر لاستخراج الإيميلات والجوالات وروابط التواصل.\n"
            "المنتجات: ينتقل إلى التصنيفات أو صفحة المتجر ويحاول استخراج المنتجات."
        )).grid(row=1, column=2, padx=3, pady=5)

        run_frame = ttk.LabelFrame(self.settings_tab, text="أداء التشغيل والتأخيرات")
        run_frame.grid(row=2, column=0, sticky="ew", padx=10, pady=10)
        for i in range(4):
            run_frame.columnconfigure(i, weight=1)

        self._small_field(run_frame, 0, 0, "التوازي", self.concurrency_var, "عدد الصفحات التي تعمل في نفس الوقت داخل نفس جلسة المتصفح. ابدأ بـ 1 أو 2 ثم ارفع بحذر.")
        self._small_field(run_frame, 0, 2, "Timeout ms", self.timeout_var, "مهلة فتح الصفحة بالملي ثانية.")
        self._small_field(run_frame, 1, 0, "حفظ كل", self.save_every_var, "يحفظ النتائج مرحليًا بعد هذا العدد من المتاجر.")
        self._small_field(run_frame, 1, 2, "Restart كل", self.restart_every_var, "إعادة تشغيل الجلسة دوريًا بعد عدد معين من المتاجر. ضع 0 لتعطيلها.")
        self._small_field(run_frame, 2, 0, "أقل تأخير", self.min_delay_var, "أقل تأخير عشوائي بين خطوات التصفح.")
        self._small_field(run_frame, 2, 2, "أعلى تأخير", self.max_delay_var, "أعلى تأخير عشوائي بين خطوات التصفح.")
        self._small_field(run_frame, 3, 0, "انتظار تحقق بشري", self.human_wait_var, "عند اكتشاف صفحة تحقق بشري يتم إغلاق الجلسة والانتظار هذه المدة ثم فتحها مرة أخرى بدون تجاوز التحقق.")
        self._small_field(run_frame, 3, 2, "محاولات إعادة", self.human_retries_var, "عدد مرات إعادة المحاولة بعد صفحة تحقق بشري.")
        ttk.Checkbutton(run_frame, text="Headless", variable=self.headless_var).grid(row=4, column=1, sticky="w", padx=6, pady=5)
        ttk.Button(run_frame, text="شرح Headless", command=lambda: self._help(
            "Headless", "عند تفعيله يعمل المتصفح بدون نافذة. عند استخدام جلسة فعلية محلية يفضل تركه غير مفعل."
        )).grid(row=4, column=2, sticky="w", padx=6, pady=5)
        self._build_quick_guide()

    def _small_field(self, parent, row, col, label, var, help_text):
        cell = ttk.Frame(parent)
        cell.grid(row=row, column=col, columnspan=2, sticky="w", padx=6, pady=5)
        ttk.Label(cell, text=label).pack(side="left", padx=(0, 6))
        ttk.Entry(cell, textvariable=var, width=12).pack(side="left", padx=(0, 4))
        ttk.Button(cell, text="؟", width=3, command=lambda: self._help(label, help_text)).pack(side="left")

    def _build_quick_guide(self):
        guide = ttk.LabelFrame(
            self.settings_tab,
            text="كتيب التشغيل السريع والشرح التفصيلي"
        )
        guide.grid(row=3, column=0, sticky="nsew", padx=10, pady=(0, 10))
        guide.columnconfigure(1, weight=1)
        guide.rowconfigure(1, weight=1)

        topics_bar = ttk.Frame(guide)
        topics_bar.grid(row=0, column=0, columnspan=2, sticky="ew", padx=6, pady=6)

        topics = [
            ("خطوات التشغيل", "workflow"),
            ("الإعدادات المقترحة", "settings"),
            ("اختيار المودات", "modes"),
            ("حالات Status", "statuses"),
            ("مود المنتجات و Selectors", "products"),
            ("ترتيب Excel", "sort"),
        ]

        for title, key in topics:
            ttk.Button(
                topics_bar,
                text=title,
                command=lambda k=key: self._set_guide_topic(k)
            ).pack(side="left", padx=3, pady=2)

        quick_box = ttk.LabelFrame(guide, text="تنقل سريع")
        quick_box.grid(row=1, column=0, sticky="nsw", padx=6, pady=6)

        ttk.Button(
            quick_box,
            text="1) فتح صفحة عرض الشيت",
            command=lambda: self.notebook.select(self.sheet_tab)
        ).pack(fill="x", padx=6, pady=4)

        ttk.Button(
            quick_box,
            text="2) قراءة الشيت الآن",
            command=self.load_sheet_preview
        ).pack(fill="x", padx=6, pady=4)

        ttk.Button(
            quick_box,
            text="3) تطبيق العمود المكتشف",
            command=self._apply_detected_columns
        ).pack(fill="x", padx=6, pady=4)

        ttk.Button(
            quick_box,
            text="4) فتح صفحة Selectors",
            command=lambda: self.notebook.select(self.advanced_tab)
        ).pack(fill="x", padx=6, pady=4)

        ttk.Button(
            quick_box,
            text="5) فتح صفحة التشغيل",
            command=lambda: self.notebook.select(self.run_tab)
        ).pack(fill="x", padx=6, pady=4)

        ttk.Button(
            quick_box,
            text="6) فتح ترتيب Excel",
            command=lambda: self.notebook.select(self.sort_tab)
        ).pack(fill="x", padx=6, pady=4)

        text_frame = ttk.Frame(guide)
        text_frame.grid(row=1, column=1, sticky="nsew", padx=6, pady=6)
        text_frame.columnconfigure(0, weight=1)
        text_frame.rowconfigure(0, weight=1)

        self.guide_text = tk.Text(
            text_frame,
            height=12,
            wrap="word",
            state="disabled",
            font=("Segoe UI", 10),
            bg="#fbfbfb",
            relief="solid",
            borderwidth=1
        )
        self.guide_text.grid(row=0, column=0, sticky="nsew")

        yscroll = ttk.Scrollbar(text_frame, orient="vertical", command=self.guide_text.yview)
        self.guide_text.configure(yscrollcommand=yscroll.set)
        yscroll.grid(row=0, column=1, sticky="ns")

        self._set_guide_topic("workflow")

    def _set_guide_topic(self, topic: str):
        if not self.guide_text:
            return

        texts = self._guide_texts()
        content = texts.get(topic, texts["workflow"])

        self.guide_text.configure(state="normal")
        self.guide_text.delete("1.0", "end")
        self.guide_text.insert("1.0", content.strip())
        self.guide_text.configure(state="disabled")

    def _guide_texts(self) -> dict[str, str]:
        return {
            "workflow": """
    خطوات التشغيل المقترحة

    1) من صفحة الإعدادات:
    - اختر ملف Excel المصدر.
    - اختر ملف Excel الناتج أو اترك القيمة الافتراضية.
    - تأكد من مسار المتصفح exe.
    - تأكد من مجلد User Data واسم Profile.
    - حدد المودات المطلوبة: فحص الحالة، بيانات التواصل، المنتجات.

    2) انتقل إلى صفحة عرض الشيت:
    - اضغط قراءة وعرض الشيت.
    - سيتم عرض أول صفوف من الملف داخل البرنامج.
    - سيتم تحديث قائمة الأعمدة تلقائيًا.

    3) ارجع إلى صفحة الإعدادات:
    - اختر عمود الروابط الصحيح من خانة عمود الروابط.
    - غالبًا يكون اسمه link أو url أو رابط.
    - لا تبدأ التشغيل قبل التأكد أن العمود المختار يحتوي روابط المتاجر فعلًا.

    4) راجع إعدادات الأداء:
    - التوازي يحدد عدد المتاجر التي يتم فتحها في نفس الوقت.
    - أقل وأعلى تأخير يحددان التأخير العشوائي بين الخطوات.
    - Timeout يحدد أقصى مدة انتظار لفتح الصفحة.
    - حفظ كل يحدد عدد النتائج قبل الحفظ المرحلي.

    5) لو ستستخدم مود المنتجات:
    - افتح صفحة Selectors والتصنيفات.
    - حدد عمود روابط التصنيفات إن كان موجودًا في الشيت.
    - اترك Selectors فارغة للاكتشاف التلقائي.
    - أو اكتب CSS Selectors مخصصة لو القالب لا يتم التعرف عليه تلقائيًا.

    6) انتقل إلى صفحة التشغيل والتقرير:
    - اضغط بدء التشغيل.
    - تابع النتائج من الجدول.
    - تابع التقدم من الشريط السفلي: منجز، المتبقي، الإجمالي.
    - بعد الانتهاء سيظهر تقرير مختصر ويتم حفظ ملف النتائج.
    """,

            "settings": """
    شرح الإعدادات والقيم المقترحة

    ملف Excel المصدر:
    الملف الذي يحتوي روابط المتاجر. يجب أن يحتوي عمودًا واضحًا للروابط مثل link أو url أو رابط.

    ملف Excel الناتج:
    الملف الذي سيتم حفظ النتائج فيه. يفضل عدم فتحه في Excel أثناء التشغيل حتى لا يحدث خطأ PermissionError.

    مسار المتصفح exe:
    مسار Brave أو Chrome. مثال:
    C:\\Program Files\\BraveSoftware\\Brave-Browser\\Application\\brave.exe

    مجلد User Data:
    مجلد بيانات المستخدم للمتصفح. الأفضل تقنيًا أن يكون مجلد User Data العام، واسم البروفايل يكون في خانة Profile.
    لكن إن كان مشروعك مضبوط حاليًا على User Data\\Default ويعمل، يمكن تركه كما هو.

    اسم Profile:
    غالبًا Default أو Profile 1. هذا يحدد الجلسة الفعلية التي سيتم تشغيل المتصفح بها.

    التوازي:
    - 1: أفضل قيمة للاختبار أو عند وجود تحقق بشري.
    - 2: قيمة مناسبة كبداية للتشغيل الفعلي.
    - 3 إلى 5: استخدمها فقط بعد التأكد أن الجهاز والاتصال والمتصفح يتحملون.

    أقل وأعلى تأخير:
    القيمة 1.0 إلى 3.0 مناسبة لفحص الحالة فقط.
    لو مود المنتجات مفعل، يفضل 2.0 إلى 5.0 لأن البرنامج يتنقل داخل التصنيفات والمنتجات.

    Timeout ms:
    - 20000 مناسب للمواقع السريعة.
    - 30000 قيمة متوازنة.
    - 60000 للمواقع الثقيلة أو الاتصال الضعيف.

    حفظ كل:
    يفضل 5 أو 10. هذا يقلل فقدان البيانات لو توقف البرنامج.

    Restart كل:
    يفضل 50 إلى 150. ضع 0 لتعطيل إعادة التشغيل الدورية.

    انتظار تحقق بشري:
    مدة الانتظار قبل إعادة فتح الجلسة عند اكتشاف صفحة تحقق بشري.

    محاولات إعادة:
    لو التوازي أكبر من 1 يفضل جعلها 0.
    لو التوازي 1 يمكن جعلها 1 أو 2.

    Headless:
    اتركه غير مفعل عند استخدام جلسة متصفح فعلية محلية.
    """,

            "modes": """
    كيفية اختيار المودات

    1) فحص الحالة
    استخدمه عندما تريد معرفة هل المتجر يعمل أم لا.
    هذا هو أسرع مود، ومناسب لاختبار ملف كبير من الروابط.

    2) بيانات التواصل
    استخدمه عندما تريد استخراج:
    - الإيميلات.
    - أرقام الجوال أو الهاتف.
    - روابط واتساب.
    - روابط انستغرام، تويتر/X، تيك توك، سناب، فيسبوك وغيرها.
    يعتمد غالبًا على الهيدر والفوتر ومحتوى الصفحة الرئيسية.

    3) المنتجات
    استخدمه عندما تريد استخراج منتجات المتجر.
    هذا هو أبطأ مود لأنه قد يفتح صفحات التصنيفات ويتنقل بينها ويحلل كروت المنتجات.

    اختيار المود المناسب:
    - لو هدفك معرفة المتاجر الشغالة فقط: فعّل فحص الحالة فقط.
    - لو هدفك بناء قائمة تواصل: فعّل فحص الحالة + بيانات التواصل.
    - لو هدفك تحليل المنتجات: فعّل الثلاثة، أو المنتجات مع الحالة.
    - عند اختبار مشروع جديد، ابدأ بفحص الحالة فقط ثم أضف باقي المودات تدريجيًا.

    ملاحظة:
    كلما زادت المودات، زاد وقت التشغيل.
    """,

            "statuses": """
    كيف يتم تحديد Status لكل موقع

    البرنامج لا يحكم على الموقع من عامل واحد فقط، بل يجمع عدة إشارات:

    1) هل الصفحة فتحت بنجاح؟
    يتم محاولة فتح الرابط داخل المتصفح بجلسة فعلية.
    لو حدث خطأ في الفتح يتم تسجيل السبب في Reason أو HTTP/Error.

    2) هل ظهرت إشارات متجر؟
    مثل:
    - header أو footer.
    - كروت منتجات.
    - زر السلة.
    - مربع البحث.
    - عناصر Salla أو Zid المعروفة.
    لو ظهرت هذه الإشارات غالبًا يتم اعتبار المتجر يعمل.

    3) هل ظهرت إشارات فشل؟
    مثل:
    - تحت الصيانة.
    - المتجر موقوف.
    - غير متاح.
    - store unavailable.
    - maintenance.
    - HTTP 404 أو 500.
    وقتها يتم اعتبار المتجر لا يعمل.

    4) هل ظهرت صفحة تحقق بشري؟
    مثل:
    - verify you are human.
    - checking your browser.
    - security check.
    - challenge.
    وقتها يتم تسجيل الحالة تحقق بشري.
    البرنامج لا يتجاوز التحقق، فقط يغلق الجلسة وينتظر ثم يعيد المحاولة حسب الإعدادات.

    حالات Status الموجودة:

    قيد الفحص:
    حالة داخلية مؤقتة أثناء معالجة الصف.

    يعمل:
    تم فتح الصفحة وظهرت إشارات متجر واضحة.

    لا يعمل:
    الصفحة فتحت لكن ظهرت إشارات توقف أو صيانة أو HTTP خطأ.

    غير مؤكد:
    الصفحة فتحت لكن لم تظهر إشارات كافية تؤكد أنها متجر يعمل، ولم تظهر إشارات فشل واضحة.

    تحقق بشري:
    تم اكتشاف صفحة تحقق أو حماية. لا يتم تجاوزها.

    خطأ:
    حدث خطأ أثناء الفتح أو المعالجة مثل:
    - رابط فارغ.
    - Timeout.
    - DNS.
    - SSL/Certificate.
    - مشكلة اتصال بالخادم.
    - خطأ غير متوقع أثناء فتح الصفحة.

    عمود Reason في النتائج يشرح السبب التفصيلي للحالة.
    """,

            "products": """
    شرح استخدام مود المنتجات وصفحة Selectors

    مود المنتجات له طريقتان:

    الطريقة الأولى: اكتشاف تلقائي
    اترك Selectors JSON كما هو والقيم فارغة.
    البرنامج سيحاول التعرف على:
    - روابط التصنيفات.
    - كروت المنتجات.
    - اسم المنتج.
    - السعر.
    - رابط المنتج.
    - صورة المنتج.

    هذه الطريقة مناسبة كبداية، لكنها قد لا تعمل مع كل القوالب.

    الطريقة الثانية: Selectors مخصصة
    استخدمها عندما يكون القالب مختلفًا أو الاكتشاف التلقائي لا يجد المنتجات.

    شرح مفاتيح Selectors JSON:

    category_links:
    CSS selector لروابط التصنيفات.
    مثال:
    nav a[href*="/categories"]

    product_card:
    CSS selector لكارت المنتج الكامل.
    مثال:
    .s-product-card-entry

    product_name:
    CSS selector لاسم المنتج داخل الكارت.
    مثال:
    .s-product-card-content-title

    product_price:
    CSS selector للسعر داخل الكارت.
    مثال:
    .s-product-card-price

    product_url:
    CSS selector لرابط المنتج داخل الكارت.
    غالبًا يكون:
    a

    product_image:
    CSS selector لصورة المنتج داخل الكارت.
    مثال:
    img

    عمود روابط التصنيفات:
    لو لديك في الشيت عمود يحتوي روابط تصنيفات جاهزة، اختره من صفحة Selectors.
    يمكن وضع أكثر من رابط في نفس الخلية مفصولًا بسطر جديد أو ; أو |.

    متى أستخدم عمود التصنيفات؟
    - عندما تريد دقة أعلى.
    - عندما لا يستطيع البرنامج اكتشاف التصنيفات تلقائيًا.
    - عندما تريد سحب أقسام معينة فقط.

    إعدادات حدود المنتجات:
    أقصى تصنيفات:
    عدد التصنيفات التي سيتم زيارتها لكل متجر.

    صفحات لكل تصنيف:
    عدد صفحات pagination داخل كل تصنيف.

    أقصى منتجات:
    الحد الأقصى للمنتجات المحفوظة لكل متجر.

    اقتراح عملي:
    ابدأ بـ:
    - أقصى تصنيفات: 3
    - صفحات لكل تصنيف: 1
    - أقصى منتجات: 20
    بعد نجاح التجربة، ارفع القيم تدريجيًا.
    """,

            "sort": """
    شرح صفحة ترتيب Excel

    هذه الصفحة مستقلة عن عملية السحب.
    تستخدمها عندما تريد ترتيب أي ملف Excel حسب عمود معين.

    طريقة الاستخدام:

    1) اختر ملف Excel للترتيب.
    هذا هو الملف الذي تريد إعادة ترتيب صفوفه.

    2) اختر ملف الخرج المرتب.
    هذا ملف جديد سيتم حفظ النسخة المرتبة فيه.
    لا يفضل الكتابة فوق الملف الأصلي.

    3) اضغط تحميل Sheets.
    سيتم عرض صفحات الملف.

    4) اختر Sheet المطلوب.
    مثال:
    Stores أو Products أو أي صفحة أخرى.

    5) اضغط تحميل الأعمدة.
    سيتم عرض أسماء الأعمدة الموجودة في أول صف.

    6) اختر عمود الترتيب.
    مثال:
    Status أو Store أو Products أو Price.

    7) اختر نوع الترتيب:
    - رقمي: استخدمها لو العمود يحتوي أرقامًا مثل عدد المنتجات أو السعر.
    - تنازلي: لترتيب من الأكبر للأصغر أو من Z إلى A.
    - بدون تنازلي: ترتيب تصاعدي من الأصغر للأكبر أو من A إلى Z.

    8) اضغط تنفيذ الترتيب.
    سيتم حفظ ملف جديد مرتب في مسار الخرج.
    """
        }

    def _build_advanced_tab(self):
        self.advanced_tab.columnconfigure(0, weight=1)
        self.advanced_tab.rowconfigure(2, weight=1)

        cat_frame = ttk.LabelFrame(self.advanced_tab, text="التصنيفات والحدود")
        cat_frame.grid(row=0, column=0, sticky="ew", padx=10, pady=10)
        cat_frame.columnconfigure(1, weight=1)

        ttk.Label(cat_frame, text="عمود روابط التصنيفات").grid(row=0, column=0, sticky="w", padx=6, pady=5)
        self.category_column_combo = ttk.Combobox(cat_frame, textvariable=self.category_column_var, values=[], state="normal")
        self.category_column_combo.grid(row=0, column=1, sticky="ew", padx=6, pady=5)
        ttk.Button(cat_frame, text="؟", width=3, command=lambda: self._help(
            "عمود روابط التصنيفات",
            "اختياري. لو لديك في الشيت عمود يحتوي روابط التصنيفات، اختره هنا. يمكن وضع أكثر من رابط في الخانة مفصولًا بسطر أو ; أو |."
        )).grid(row=0, column=2, padx=3, pady=5)

        limits = ttk.Frame(cat_frame)
        limits.grid(row=1, column=0, columnspan=3, sticky="ew")
        for i in range(6):
            limits.columnconfigure(i, weight=1)
        self._small_field(limits, 0, 0, "أقصى تصنيفات", self.max_categories_var, "أقصى عدد تصنيفات يتم زيارتها لكل متجر.")
        self._small_field(limits, 0, 2, "صفحات لكل تصنيف", self.max_category_pages_var, "أقصى عدد صفحات Pagination يتم زيارتها داخل التصنيف الواحد.")
        self._small_field(limits, 0, 4, "أقصى منتجات", self.max_products_var, "أقصى عدد منتجات يتم حفظه لكل متجر.")

        selector_frame = ttk.LabelFrame(self.advanced_tab, text="Selectors JSON")
        selector_frame.grid(row=2, column=0, sticky="nsew", padx=10, pady=10)
        selector_frame.columnconfigure(0, weight=1)
        selector_frame.rowconfigure(0, weight=1)

        self.selectors_text = tk.Text(selector_frame, height=16, wrap="none")
        self.selectors_text.insert("1.0", DEFAULT_SELECTORS_JSON)
        self.selectors_text.grid(row=0, column=0, sticky="nsew", padx=6, pady=6)

        yscroll = ttk.Scrollbar(selector_frame, orient="vertical", command=self.selectors_text.yview)
        xscroll = ttk.Scrollbar(selector_frame, orient="horizontal", command=self.selectors_text.xview)
        self.selectors_text.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        yscroll.grid(row=0, column=1, sticky="ns")
        xscroll.grid(row=1, column=0, sticky="ew")

        btns = ttk.Frame(selector_frame)
        btns.grid(row=2, column=0, sticky="w", padx=6, pady=6)
        ttk.Button(btns, text="شرح Selectors", command=lambda: self._help(
            "Selectors JSON",
            "اترك القيم فارغة للاكتشاف التلقائي.\n"
            "category_links: selector لروابط التصنيفات.\n"
            "product_card: selector لكارت المنتج.\n"
            "product_name/product_price/product_url/product_image: selectors داخل كارت المنتج."
        )).pack(side="left", padx=4)
        ttk.Button(btns, text="استعادة الافتراضي", command=self._reset_selectors).pack(side="left", padx=4)
        ttk.Button(btns, text="اختبار JSON", command=self._validate_selectors_json).pack(side="left", padx=4)

    def _build_sheet_tab(self):
        self.sheet_tab.columnconfigure(0, weight=1)
        self.sheet_tab.rowconfigure(1, weight=1)

        toolbar = ttk.Frame(self.sheet_tab)
        toolbar.grid(row=0, column=0, sticky="ew", padx=10, pady=8)
        ttk.Button(toolbar, text="قراءة وعرض الشيت", command=self.load_sheet_preview).pack(side="left", padx=4)
        ttk.Button(toolbar, text="تطبيق العمود المكتشف", command=self._apply_detected_columns).pack(side="left", padx=4)
        ttk.Button(toolbar, text="شرح", command=lambda: self._help(
            "عرض الشيت",
            "يعرض أول 100 صف من ملف المصدر، ويحدّث قائمة الأعمدة لاختيار عمود الروابط وعمود التصنيفات."
        )).pack(side="left", padx=4)

        self.preview_tree = ttk.Treeview(self.sheet_tab, show="headings")
        self.preview_tree.grid(row=1, column=0, sticky="nsew", padx=10, pady=8)
        self._attach_tree_scrollbars(self.sheet_tab, self.preview_tree, 1, 0)

    def _build_run_tab(self):
        self.run_tab.columnconfigure(0, weight=1)
        self.run_tab.rowconfigure(1, weight=2)
        self.run_tab.rowconfigure(4, weight=1)

        toolbar = ttk.Frame(self.run_tab)
        toolbar.grid(row=0, column=0, sticky="ew", padx=10, pady=8)
        self.start_button = ttk.Button(toolbar, text="بدء التشغيل", command=self.start_scraping)
        self.start_button.pack(side="left", padx=4)
        self.stop_button = ttk.Button(toolbar, text="إيقاف", command=self.stop_scraping, state="disabled")
        self.stop_button.pack(side="left", padx=4)
        ttk.Button(toolbar, text="فتح ملف النتائج", command=self._open_output_location).pack(side="left", padx=4)
        ttk.Button(toolbar, text="شرح التشغيل", command=lambda: self._help(
            "التشغيل",
            "ابدأ بعد اختيار ملف المصدر وضبط عمود الروابط. النتائج تظهر هنا ويتم حفظها في ملف Excel الناتج."
        )).pack(side="left", padx=4)

        self.result_tree = ttk.Treeview(self.run_tab, columns=("idx", "url", "platform", "name", "status", "reason", "products", "emails", "phones"), show="headings")
        headings = {
            "idx": "#", "url": "URL", "platform": "Platform", "name": "Store",
            "status": "Status", "reason": "Reason", "products": "Products",
            "emails": "Emails", "phones": "Phones",
        }
        widths = {"idx": 50, "url": 240, "platform": 80, "name": 180, "status": 110, "reason": 260, "products": 80, "emails": 180, "phones": 150}
        for col, heading in headings.items():
            self.result_tree.heading(col, text=heading)
            self.result_tree.column(col, width=widths[col], stretch=True)
        self.result_tree.grid(row=1, column=0, sticky="nsew", padx=10, pady=8)
        self._attach_tree_scrollbars(self.run_tab, self.result_tree, 1, 0)

        ttk.Label(self.run_tab, text="Logs").grid(row=3, column=0, sticky="w", padx=10)
        self.log_text = tk.Text(self.run_tab, height=9, wrap="word", state="disabled")
        self.log_text.grid(row=4, column=0, sticky="nsew", padx=10, pady=8)

    def _build_sort_tab(self):
        self.sort_tab.columnconfigure(0, weight=1)
        frame = ttk.LabelFrame(self.sort_tab, text="ترتيب ملف Excel مستقل")
        frame.grid(row=0, column=0, sticky="ew", padx=10, pady=10)
        self._field(frame, 0, "ملف Excel للترتيب", self.sort_input_var,
                    "اختر ملف Excel تريد ترتيبه حسب عمود معين.", "open_file")
        self._field(frame, 1, "ملف الخرج المرتب", self.sort_output_var,
                    "الملف الجديد بعد الترتيب.", "save_file")

        ttk.Label(frame, text="Sheet").grid(row=2, column=0, sticky="w", padx=6, pady=5)
        self.sort_sheet_combo = ttk.Combobox(frame, textvariable=self.sort_sheet_var, values=[], state="normal")
        self.sort_sheet_combo.grid(row=2, column=1, sticky="ew", padx=6, pady=5)
        self.sort_sheet_combo.bind("<<ComboboxSelected>>", lambda e: self._load_sort_columns())
        ttk.Button(frame, text="تحميل Sheets", command=self._load_sort_sheets).grid(row=2, column=2, padx=3, pady=5)
        ttk.Button(frame, text="؟", width=3, command=lambda: self._help("Sheet", "اختر اسم الصفحة داخل ملف Excel.")).grid(row=2, column=3, padx=3, pady=5)

        ttk.Label(frame, text="عمود الترتيب").grid(row=3, column=0, sticky="w", padx=6, pady=5)
        self.sort_column_combo = ttk.Combobox(frame, textvariable=self.sort_column_var, values=[], state="normal")
        self.sort_column_combo.grid(row=3, column=1, sticky="ew", padx=6, pady=5)
        ttk.Button(frame, text="تحميل الأعمدة", command=self._load_sort_columns).grid(row=3, column=2, padx=3, pady=5)
        ttk.Button(frame, text="؟", width=3, command=lambda: self._help("عمود الترتيب", "العمود الذي سيتم ترتيب الصفوف بناء عليه.")).grid(row=3, column=3, padx=3, pady=5)

        options = ttk.Frame(frame)
        options.grid(row=4, column=1, sticky="w", padx=6, pady=5)
        ttk.Checkbutton(options, text="رقمي", variable=self.sort_numeric_var).pack(side="left", padx=8)
        ttk.Checkbutton(options, text="تنازلي", variable=self.sort_desc_var).pack(side="left", padx=8)

        ttk.Button(frame, text="تنفيذ الترتيب", command=self._run_sort).grid(row=5, column=1, sticky="w", padx=6, pady=10)
        ttk.Button(frame, text="شرح أداة الترتيب", command=lambda: self._help(
            "ترتيب Excel",
            "هذه الأداة منفصلة عن السحب. تختار ملفًا وشيت وعمودًا، ثم تحفظ نسخة مرتبة أبجديًا أو رقميًا تصاعديًا/تنازليًا."
        )).grid(row=5, column=2, sticky="w", padx=6, pady=10)

    def _build_footer(self):
        footer = ttk.Frame(self)
        footer.grid(row=1, column=0, sticky="ew", padx=8, pady=(0, 8))
        footer.columnconfigure(0, weight=1)

        self.progress = ttk.Progressbar(footer, mode="determinate")
        self.progress.grid(row=0, column=0, sticky="ew", padx=6, pady=4)

        stats = ttk.Frame(footer)
        stats.grid(row=0, column=1, sticky="e")
        ttk.Label(stats, textvariable=self.status_var).pack(side="left", padx=8)
        ttk.Label(stats, text="منجز:").pack(side="left")
        ttk.Label(stats, textvariable=self.done_var).pack(side="left", padx=3)
        ttk.Label(stats, text="المتبقي:").pack(side="left", padx=(12, 0))
        ttk.Label(stats, textvariable=self.remaining_var).pack(side="left", padx=3)
        ttk.Label(stats, text="الإجمالي:").pack(side="left", padx=(12, 0))
        ttk.Label(stats, textvariable=self.total_var).pack(side="left", padx=3)

    def _attach_tree_scrollbars(self, parent, tree, row, col):
        yscroll = ttk.Scrollbar(parent, orient="vertical", command=tree.yview)
        xscroll = ttk.Scrollbar(parent, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        yscroll.grid(row=row, column=col + 1, sticky="ns", pady=8)
        xscroll.grid(row=row + 1, column=col, sticky="ew", padx=10)

    def _browse_open_file(self, variable: tk.StringVar):
        path = filedialog.askopenfilename(filetypes=[("Excel / Executable", "*.xlsx *.xlsm *.exe"), ("All files", "*.*")])
        if path:
            variable.set(path)

    def _browse_save_file(self, variable: tk.StringVar):
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx"), ("All files", "*.*")])
        if path:
            variable.set(path)

    def _browse_dir(self, variable: tk.StringVar):
        path = filedialog.askdirectory()
        if path:
            variable.set(path)

    def _reset_selectors(self):
        if self.selectors_text:
            self.selectors_text.delete("1.0", "end")
            self.selectors_text.insert("1.0", DEFAULT_SELECTORS_JSON)

    def _validate_selectors_json(self):
        try:
            parse_selectors_json(self.selectors_text.get("1.0", "end") if self.selectors_text else "")
            messagebox.showinfo("JSON", "Selectors JSON صحيح.")
        except Exception as exc:
            messagebox.showerror("JSON Error", str(exc))

    def load_sheet_preview(self):
        path = self.input_excel_var.get().strip()
        if not path or not os.path.exists(path):
            messagebox.showerror("خطأ", "اختر ملف Excel مصدر صحيح.")
            return
        try:
            headers, rows, detected = preview_rows(path)
            self.headers = headers
            self.link_column_combo.configure(values=headers)
            self.category_column_combo.configure(values=[""] + headers)
            self.link_column_var.set(detected)
            self._populate_tree(self.preview_tree, headers, rows)
            self.status_var.set("تم عرض الشيت")
            self.total_var.set(str(max(0, len(rows))))
            self._log(f"تم قراءة الشيت. العمود المكتشف للروابط: {detected}")
        except Exception as exc:
            messagebox.showerror("خطأ قراءة الشيت", str(exc))

    def _apply_detected_columns(self):
        if not self.headers:
            self.load_sheet_preview()
            return
        detected = detect_link_column(self.headers)
        self.link_column_var.set(detected)
        self.link_column_combo.configure(values=self.headers)
        self.category_column_combo.configure(values=[""] + self.headers)
        self._log(f"تم تطبيق العمود المكتشف: {detected}")

    def _populate_tree(self, tree: ttk.Treeview, headers: list[str], rows: list[list]):
        for item in tree.get_children():
            tree.delete(item)
        tree.configure(columns=headers)
        for header in headers:
            tree.heading(header, text=header)
            tree.column(header, width=max(100, min(220, len(header) * 12)), stretch=True)
        for row in rows:
            values = list(row) + [""] * max(0, len(headers) - len(row))
            tree.insert("", "end", values=values[: len(headers)])

    def build_settings(self) -> ScrapeSettings:
        input_excel = self.input_excel_var.get().strip()
        if not input_excel or not os.path.exists(input_excel):
            raise ValueError("ملف Excel المصدر غير صحيح.")

        output_excel = self.output_excel_var.get().strip()
        if not output_excel:
            output_excel = str(Path(input_excel).with_name("salla_zid_results.xlsx"))

        modes = []
        if self.mode_status_var.get():
            modes.append("status")
        if self.mode_contacts_var.get():
            modes.append("contacts")
        if self.mode_products_var.get():
            modes.append("products")
        if not modes:
            raise ValueError("اختر مود واحد على الأقل.")

        selectors = parse_selectors_json(self.selectors_text.get("1.0", "end") if self.selectors_text else "")

        return ScrapeSettings(
            input_excel=input_excel,
            output_excel=output_excel,
            link_column=self.link_column_var.get().strip() or "link",
            browser_exe=self.browser_exe_var.get().strip(),
            user_data_dir=self.user_data_dir_var.get().strip(),
            profile_dir=self.profile_dir_var.get().strip() or "Default",
            headless=self.headless_var.get(),
            modes=modes,
            concurrency=clamp_int(self.concurrency_var.get(), 2, 1, 10),
            navigation_timeout_ms=clamp_int(self.timeout_var.get(), 30000, 5000, 180000),
            save_every=clamp_int(self.save_every_var.get(), 5, 1, 1000),
            restart_every=clamp_int(self.restart_every_var.get(), 100, 0, 100000),
            min_delay_sec=clamp_float(self.min_delay_var.get(), 1.0, 0.0, 60.0),
            max_delay_sec=clamp_float(self.max_delay_var.get(), 3.0, 0.0, 120.0),
            human_verification_wait_sec=clamp_int(self.human_wait_var.get(), 60, 1, 3600),
            human_verification_retries=clamp_int(self.human_retries_var.get(), 1, 0, 10),
            max_categories_per_store=clamp_int(self.max_categories_var.get(), 20, 1, 500),
            max_category_pages=clamp_int(self.max_category_pages_var.get(), 3, 1, 50),
            max_products_per_store=clamp_int(self.max_products_var.get(), 200, 1, 10000),
            category_urls_column=self.category_column_var.get().strip(),
            selectors=selectors,
        )

    def start_scraping(self):
        if self.worker_thread and self.worker_thread.is_alive():
            messagebox.showwarning("يعمل بالفعل", "هناك عملية تشغيل قائمة.")
            return
        try:
            settings = self.build_settings()
        except Exception as exc:
            messagebox.showerror("خطأ في الإعدادات", str(exc))
            return

        self.results = []
        self.stop_event.clear()
        self.result_tree.delete(*self.result_tree.get_children())
        self.start_button.configure(state="disabled")
        self.stop_button.configure(state="normal")
        self.progress.configure(value=0, maximum=1)
        self.done_var.set("0")
        self.remaining_var.set("0")
        self.total_var.set("0")
        self.status_var.set("يعمل")
        self._log("بدأ التشغيل...")

        def emit(event: dict):
            self.ui_queue.put(event)

        def target():
            try:
                asyncio.run(run_scraper(settings, emit, self.stop_event.is_set))
            except Exception:
                self.ui_queue.put({"type": "fatal", "error": traceback.format_exc()})

        self.worker_thread = threading.Thread(target=target, daemon=True)
        self.worker_thread.start()
        self.notebook.select(self.run_tab)

    def stop_scraping(self):
        self.stop_event.set()
        self.status_var.set("إيقاف مطلوب")
        self._log("تم إرسال طلب إيقاف. سيتم التوقف بعد انتهاء الخطوة الحالية.", "WARN")

    def _poll_queue(self):
        try:
            while True:
                event = self.ui_queue.get_nowait()
                self._handle_event(event)
        except queue.Empty:
            pass
        self.after(200, self._poll_queue)

    def _handle_event(self, event: dict):
        kind = event.get("type")
        if kind == "log":
            self._log(event.get("message", ""), event.get("level", "INFO"))
        elif kind == "headers":
            headers = event.get("headers") or []
            self.headers = headers
            self.link_column_combo.configure(values=headers)
            self.category_column_combo.configure(values=[""] + headers)
            self.link_column_var.set(event.get("link_column") or self.link_column_var.get())
        elif kind == "progress":
            done = int(event.get("done", 0))
            total = int(event.get("total", 0))
            remaining = int(event.get("remaining", max(total - done, 0)))
            self.progress.configure(maximum=max(total, 1), value=done)
            self.done_var.set(str(done))
            self.total_var.set(str(total))
            self.remaining_var.set(str(remaining))
            self.status_var.set("يعمل" if remaining else "انتهى")
        elif kind == "result":
            result = event.get("result")
            if result:
                self.results.append(result)
                self._add_result_row(result)
        elif kind == "human_wait":
            self.status_var.set(f"انتظار تحقق بشري: {event.get('remaining')}s")
        elif kind == "done":
            self.start_button.configure(state="normal")
            self.stop_button.configure(state="disabled")
            output = event.get("output_excel", "")
            self.status_var.set("تم الانتهاء")
            self._log(f"اكتمل التشغيل. ملف النتائج: {output}")
            self._show_summary()
        elif kind == "fatal":
            self.start_button.configure(state="normal")
            self.stop_button.configure(state="disabled")
            self.status_var.set("خطأ")
            self._log(event.get("error", ""), "ERROR")
            messagebox.showerror("خطأ قاتل", event.get("error", ""))
        else:
            self._log(str(event), "DEBUG")

    def _add_result_row(self, result: StoreResult):
        self.result_tree.insert("", "end", values=(
            result.index + 1,
            result.input_url,
            result.platform,
            result.store_name,
            result.status,
            result.status_reason,
            result.products_found,
            result.emails,
            result.phones,
        ))

    def _log(self, message: str, level: str = "INFO"):
        if not hasattr(self, "log_text"):
            return
        self.log_text.configure(state="normal")
        self.log_text.insert("end", f"[{level}] {message}\n")
        self.log_text.see("end")
        self.log_text.configure(state="disabled")

    def _show_summary(self):
        total = len(self.results)
        statuses: dict[str, int] = {}
        product_success = 0
        for result in self.results:
            statuses[result.status] = statuses.get(result.status, 0) + 1
            if result.products_found:
                product_success += 1
        lines = [f"إجمالي المتاجر: {total}", f"متاجر تم سحب منتجات منها: {product_success}"]
        for status, count in sorted(statuses.items()):
            lines.append(f"{status}: {count}")
        messagebox.showinfo("تقرير مختصر", "\n".join(lines))

    def _open_output_location(self):
        path = self.output_excel_var.get().strip()
        if not path:
            return
        folder = os.path.dirname(os.path.abspath(path))
        if os.path.exists(folder):
            try:
                os.startfile(folder)  # type: ignore[attr-defined]
            except Exception:
                messagebox.showinfo("المسار", folder)
        else:
            messagebox.showwarning("غير موجود", "مجلد الخرج غير موجود بعد.")

    def _load_sort_sheets(self):
        path = self.sort_input_var.get().strip()
        if not path or not os.path.exists(path):
            messagebox.showerror("خطأ", "اختر ملف Excel صحيح.")
            return
        try:
            sheets = get_sheet_names(path)
            self.sort_sheet_combo.configure(values=sheets)
            if sheets:
                self.sort_sheet_var.set(sheets[0])
                self._load_sort_columns()
        except Exception as exc:
            messagebox.showerror("خطأ", str(exc))

    def _load_sort_columns(self):
        path = self.sort_input_var.get().strip()
        sheet = self.sort_sheet_var.get().strip()
        if not path or not os.path.exists(path):
            messagebox.showerror("خطأ", "اختر ملف Excel صحيح.")
            return
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
            headers = [safe_str(c.value) or f"Column {c.column}" for c in ws[1]]
            self.sort_column_combo.configure(values=headers)
            if headers and not self.sort_column_var.get():
                self.sort_column_var.set(headers[0])
        except Exception as exc:
            messagebox.showerror("خطأ", str(exc))

    def _run_sort(self):
        try:
            sort_excel(
                self.sort_input_var.get().strip(),
                self.sort_output_var.get().strip(),
                self.sort_sheet_var.get().strip(),
                self.sort_column_var.get().strip(),
                numeric=self.sort_numeric_var.get(),
                descending=self.sort_desc_var.get(),
            )
            messagebox.showinfo("تم", f"تم حفظ الملف المرتب:\n{self.sort_output_var.get().strip()}")
        except Exception as exc:
            messagebox.showerror("خطأ الترتيب", str(exc))

def main():
    app = App()
    app.mainloop()
